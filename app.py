import streamlit as st
import pandas as pd
import yfinance as yf
import numpy as np
import requests
from io import BytesIO
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="Macro ETF Strategy Dashboard", layout="wide")

st.title("🌍 Global Macro Strategy Dashboard")

STATIC_DIR = Path("data/static")
VALUATION_PATH = STATIC_DIR / "valuation_ranks.csv"
NARRATIVE_PATH = STATIC_DIR / "narrative_ranks.csv"

# --- 1. SESSION STATE & DEFAULTS ---
DEFAULTS = {
    "w_macro": 0.20,
    "w_curr": 0.20,
    "w_val": 0.20,
    "w_nar": 0.30,
    "w_mcap": 0.10,
    "horizon": 10,
    "oil_scenario": 10.0
}

for key, val in DEFAULTS.items():
    if key not in st.session_state:
        st.session_state[key] = val

if "calc_snapshot" not in st.session_state:
    st.session_state.calc_snapshot = None

if "research_editor_version" not in st.session_state:
    st.session_state.research_editor_version = 0

# --- 2. CONFIGURATION & MAPPINGS ---
COUNTRY_CONFIG = {
    "Argentina": {"iso3": "ARG", "bis": "AR", "ticker": "ARGT", "region": "Latin America ex-Brazil/Mexico"},
    "Australia": {"iso3": "AUS", "bis": "AU", "ticker": "EWA", "region": "Australia"},
    "Austria": {"iso3": "AUT", "bis": "AT", "ticker": "EWO", "region": "Core Europe"},
    "Belgium": {"iso3": "BEL", "bis": "BE", "ticker": "EWK", "region": "Core Europe"},
    "Brazil": {"iso3": "BRA", "bis": "BR", "ticker": "EWZ", "region": "Brazil"},
    "Canada": {"iso3": "CAN", "bis": "CA", "ticker": "EWC", "region": "Canada"},
    "Chile": {"iso3": "CHL", "bis": "CL", "ticker": "ECH", "region": "Latin America ex-Brazil/Mexico"},
    "China": {"iso3": "CHN", "bis": "CN", "ticker": "MCHI", "region": "China / Hong Kong"},
    "Colombia": {"iso3": "COL", "bis": "CO", "ticker": "GXG", "region": "Latin America ex-Brazil/Mexico"},
    "Denmark": {"iso3": "DNK", "bis": "DK", "ticker": "EDEN", "region": "Nordics"},
    "Finland": {"iso3": "FIN", "bis": "FI", "ticker": "EFNL", "region": "Nordics"},
    "France": {"iso3": "FRA", "bis": "FR", "ticker": "EWQ", "region": "Core Europe"},
    "Germany": {"iso3": "DEU", "bis": "DE", "ticker": "EWG", "region": "Core Europe"},
    "Greece": {"iso3": "GRC", "bis": "GR", "ticker": "GREK", "region": "Southern Europe"},
    "Hong Kong": {"iso3": "HKG", "bis": "HK", "ticker": "EWH", "region": "China / Hong Kong"},
    "India": {"iso3": "IND", "bis": "IN", "ticker": "INDA", "region": "India"},
    "Indonesia": {"iso3": "IDN", "bis": "ID", "ticker": "EIDO", "region": "Southeast Asia"},
    "Ireland": {"iso3": "IRL", "bis": "IE", "ticker": "EIRL", "region": "Ireland"},
    "Israel": {"iso3": "ISR", "bis": "IL", "ticker": "EIS", "region": "Middle East"},
    "Italy": {"iso3": "ITA", "bis": "IT", "ticker": "EWI", "region": "Southern Europe"},
    "Japan": {"iso3": "JPN", "bis": "JP", "ticker": "EWJ", "region": "Japan"},
    "Malaysia": {"iso3": "MYS", "bis": "MY", "ticker": "EWM", "region": "Southeast Asia"},
    "Mexico": {"iso3": "MEX", "bis": "MX", "ticker": "EWW", "region": "Mexico"},
    "Netherlands": {"iso3": "NLD", "bis": "NL", "ticker": "EWN", "region": "Core Europe"},
    "Norway": {"iso3": "NOR", "bis": "NO", "ticker": "ENOR", "region": "Nordics"},
    "Philippines": {"iso3": "PHL", "bis": "PH", "ticker": "EPHE", "region": "Southeast Asia"},
    "Poland": {"iso3": "POL", "bis": "PL", "ticker": "EPOL", "region": "Eastern Europe"},
    "Singapore": {"iso3": "SGP", "bis": "SG", "ticker": "EWS", "region": "Southeast Asia"},
    "South Africa": {"iso3": "ZAF", "bis": "ZA", "ticker": "EZA", "region": "South Africa"},
    "South Korea": {"iso3": "KOR", "bis": "KR", "ticker": "EWY", "region": "South Korea"},
    "Spain": {"iso3": "ESP", "bis": "ES", "ticker": "EWP", "region": "Southern Europe"},
    "Sweden": {"iso3": "SWE", "bis": "SE", "ticker": "EWD", "region": "Nordics"},
    "Switzerland": {"iso3": "CHE", "bis": "CH", "ticker": "EWL", "region": "Core Europe"},
    "Taiwan": {"iso3": "TWN", "bis": "TW", "ticker": "EWT", "region": "Taiwan"},
    "Thailand": {"iso3": "THA", "bis": "TH", "ticker": "THD", "region": "Southeast Asia"},
    "Turkey": {"iso3": "TUR", "bis": "TR", "ticker": "TUR", "region": "Turkey"},
    "United Kingdom": {"iso3": "GBR", "bis": "GB", "ticker": "EWU", "region": "United Kingdom"},
    "United States": {"iso3": "USA", "bis": "US", "ticker": "IVV", "region": "United States"}
}

# --- 3. WIDE DATA FETCHERS ---

@st.cache_data
def fetch_imf_gdp_journey():
    iso_list = [v['iso3'] for v in COUNTRY_CONFIG.values()]
    end_year = 2025
    start_year = end_year - 10
    url = f"https://api.imf.org/external/sdmx/2.1/data/WEO/{'+'.join(iso_list)}.NGDPD.A?startPeriod={start_year}&endPeriod={end_year}"
    try:
        response = requests.get(url, headers={"Accept": "application/json"}, timeout=15)
        response.raise_for_status()
        data = response.json()
        structure = data.get('structure', {})
        country_map = {i: v['id'] for i, v in enumerate(structure.get('dimensions', {}).get('series', [])[0].get('values', []))}
        time_map = {i: v['id'] for i, v in enumerate(structure.get('dimensions', {}).get('observation', [])[0].get('values', []))}
        iso_to_name = {v['iso3']: k for k, v in COUNTRY_CONFIG.items()}
        results = []
        series_data = data.get('dataSets', [{}])[0].get('series', {})
        for key, s_val in series_data.items():
            iso = country_map.get(int(key.split(':')[0]))
            name = iso_to_name.get(iso)
            if not name: continue
            obs = s_val.get('observations', {})
            row = {"Country": name}
            vals = {}
            for t_idx, v_list in obs.items():
                yr = int(time_map.get(int(t_idx)))
                val = float(v_list[0])
                row[f"GDP_{yr}"] = val / 1e9
                vals[yr] = val
            for n in [1, 3, 5, 10]:
                v_end, v_start = vals.get(2025), vals.get(2025-n)
                row[f"GDP_CAGR_{n}Y"] = (v_end/v_start)**(1/n) - 1 if v_end and v_start and v_start > 0 else np.nan
            results.append(row)
        return pd.DataFrame(results)
    except: return pd.DataFrame()

@st.cache_data
def fetch_etf_journey(terminal_date):
    term_dt = pd.to_datetime(terminal_date)
    start_date = term_dt - pd.DateOffset(years=11)
    ytd_start = datetime(2026, 1, 1)
    results = []
    progress_bar = st.progress(0, text="Fetching ETF prices...")
    tickers = list(COUNTRY_CONFIG.items())
    for i, (name, meta) in enumerate(tickers):
        try:
            df = yf.download(meta['ticker'], start=start_date.strftime("%Y-%m-%d"), progress=False, auto_adjust=True)
            if not df.empty:
                close = df['Close'][meta['ticker']] if isinstance(df.columns, pd.MultiIndex) else df['Close']
                valid = close[close.index <= term_dt]
                if not valid.empty and len(valid) > 20:
                    p_now = float(valid.iloc[-1])
                    row = {"Country": name, "ETF Ticker": meta['ticker'], "P_Now": p_now}
                    for n in [1, 3, 5, 10]:
                        target_dt = valid.index[-1] - pd.DateOffset(years=n)
                        idx = valid.index.get_indexer([target_dt], method='nearest')[0]
                        p_start = float(valid.iloc[idx])
                        row[f"P_{n}Y"] = p_start
                        row[f"ETF_CAGR_{n}Y"] = (p_now/p_start)**(1/n) - 1 if p_start > 0 else np.nan
                    ytd_data = close[close.index >= ytd_start]
                    row["YTD_Return"] = (ytd_data.iloc[-1] / ytd_data.iloc[0]) - 1 if not ytd_data.empty else np.nan
                    results.append(row)
        except: continue
        progress_bar.progress((i + 1) / len(tickers))
    progress_bar.empty()
    return pd.DataFrame(results)

@st.cache_data
def fetch_bis_reer_journey():
    url = "https://stats.bis.org/api/v2/data/dataflow/BIS/WS_EER/1.0/M.R.B?format=csv"
    try:
        df = pd.read_csv(url)
        bis_codes = {v['bis']: k for k, v in COUNTRY_CONFIG.items()}
        df = df[df['REF_AREA'].isin(bis_codes.keys())].copy()
        df["date"] = pd.to_datetime(df["TIME_PERIOD"])
        results = []
        for code, group in df.groupby('REF_AREA'):
            group = group.sort_values("date")
            p_now = group.iloc[-1]["OBS_VALUE"]
            avg_10y = group[group["date"] >= (group["date"].max() - pd.DateOffset(years=10))]["OBS_VALUE"].mean()
            results.append({"Country": bis_codes[code], "Current_REER": p_now, "Avg_REER_10Y": avg_10y, "REER_Upside": (avg_10y - p_now) / avg_10y if avg_10y > 0 else np.nan})
        return pd.DataFrame(results)
    except: return pd.DataFrame()

@st.cache_data
def fetch_oil_impact_journey():
    """Fetch crude oil import data and map to COUNTRY_CONFIG."""
    excel_path = STATIC_DIR / "oil_imports_2024.xlsx"
    try:
        df = pd.read_excel(excel_path, sheet_name='2024 imports')
        df.columns = [c.strip() for c in df.columns]
        mappings = {"Korea, Rep.": "South Korea", "Hong Kong, China": "Hong Kong", "Other Asia, nes": "Taiwan"}
        df['Country'] = df['Country'].replace(mappings)
        cols = ['Country', 'Crude (kg)', 'Crude (mb/d)']
        df_clean = df[df['Country'].isin(COUNTRY_CONFIG.keys())][cols].copy()
        df_clean['Conversion_Factor'] = 7.33
        return df_clean
    except Exception as e:
        st.error(f"❌ Oil Impact Data Error: {e}")
        return pd.DataFrame(columns=['Country', 'Crude (kg)', 'Crude (mb/d)', 'Conversion_Factor'])


def load_research_inputs():
    return (
        pd.read_csv(VALUATION_PATH),
        pd.read_csv(NARRATIVE_PATH),
    )


def build_research_editor_inputs(val_source_df, nar_source_df):
    """Expose model-facing final ranks, not raw static-file column names."""
    countries = pd.DataFrame({"Country": list(COUNTRY_CONFIG.keys())})
    val_editor = countries.merge(val_source_df[["Country", "Average Rank"]], on="Country", how="left")
    nar_editor = countries.merge(nar_source_df[["Country", "Rank"]], on="Country", how="left")
    val_editor["Valuation Rank"] = pd.to_numeric(val_editor["Average Rank"], errors="coerce").rank(method="first", ascending=True).astype("Int64")
    nar_editor["Narrative Rank"] = pd.to_numeric(nar_editor["Rank"], errors="coerce").rank(method="first", ascending=True).astype("Int64")
    return (
        val_editor[["Country", "Valuation Rank"]],
        nar_editor[["Country", "Narrative Rank"]],
    )


def normalize_research_inputs(val_df, nar_df):
    # The app edits final model ranks. These are mapped back to the legacy
    # static-file columns because downstream workbook/export code expects them.
    val_clean = val_df[["Country", "Valuation Rank"]].rename(columns={"Valuation Rank": "Average Rank"}).copy()
    nar_clean = nar_df[["Country", "Narrative Rank"]].rename(columns={"Narrative Rank": "Rank"}).copy()
    val_clean["Country"] = val_clean["Country"].astype(str).str.strip()
    nar_clean["Country"] = nar_clean["Country"].astype(str).str.strip()
    val_clean["Average Rank"] = pd.to_numeric(val_clean["Average Rank"], errors="coerce")
    nar_clean["Rank"] = pd.to_numeric(nar_clean["Rank"], errors="coerce")
    return val_clean, nar_clean


def merge_research_inputs_for_save(saved_df, edited_df, value_col, storage_col):
    merged_df = saved_df.copy()
    edited_clean = edited_df[["Country", value_col]].rename(columns={value_col: storage_col}).copy()
    edited_clean["Country"] = edited_clean["Country"].astype(str).str.strip()
    edited_clean[storage_col] = pd.to_numeric(edited_clean[storage_col], errors="coerce")
    update_map = edited_clean.set_index("Country")[storage_col]
    if storage_col not in merged_df.columns:
        merged_df[storage_col] = np.nan
    merged_df[storage_col] = merged_df["Country"].map(update_map).combine_first(merged_df[storage_col])
    configured_countries = set(COUNTRY_CONFIG.keys())
    missing_saved = sorted(configured_countries - set(merged_df["Country"]))
    if missing_saved:
        additions = pd.DataFrame({"Country": missing_saved, storage_col: [update_map.get(country, np.nan) for country in missing_saved]})
        merged_df = pd.concat([merged_df, additions], ignore_index=True)
    return merged_df


def validate_research_inputs(val_df, nar_df):
    errors = []
    required = {
        "Valuation": (val_df, ["Country", "Valuation Rank"], "Valuation Rank"),
        "Narrative": (nar_df, ["Country", "Narrative Rank"], "Narrative Rank"),
    }
    for label, (df, columns, numeric_col) in required.items():
        missing_cols = [col for col in columns if col not in df.columns]
        if missing_cols:
            errors.append(f"{label} is missing column(s): {', '.join(missing_cols)}")
            continue
        countries = df["Country"].astype(str).str.strip()
        if countries.eq("").any() or df["Country"].isna().any():
            errors.append(f"{label} has blank country values.")
        duplicate_countries = sorted(countries[countries.duplicated()].unique())
        if duplicate_countries:
            errors.append(f"{label} has duplicate countries: {', '.join(duplicate_countries)}")
        numeric_values = pd.to_numeric(df[numeric_col], errors="coerce")
        raw_values = df[numeric_col]
        nonblank_values = raw_values.notna() & raw_values.astype(str).str.strip().ne("")
        invalid_nonblank = countries[numeric_values.isna() & nonblank_values].tolist()
        if invalid_nonblank:
            errors.append(f"{label} has non-numeric {numeric_col} values for: {', '.join(invalid_nonblank)}")
    return errors


def build_master_dataset(gdp_df, etf_df, reer_df, oil_df, mcap_df, bond_df, val_df, nar_df):
    master_df = pd.DataFrame(COUNTRY_CONFIG.keys(), columns=["Country"])
    master_df["Region"] = master_df["Country"].map(lambda x: COUNTRY_CONFIG[x]["region"])
    for d in [gdp_df, etf_df, reer_df, oil_df, mcap_df, bond_df, val_df]:
        if not d.empty:
            master_df = master_df.merge(d, on="Country", how="left")
    return master_df.merge(nar_df, on="Country", how="left")


def calculate_strategy(master_df, horizon, weights):
    required = [f'GDP_CAGR_{horizon}Y', f'ETF_CAGR_{horizon}Y', 'Mcap_USD_Bn', 'REER_Upside', 'differential with USA', 'Average Rank', 'Rank']
    calc_df = master_df.dropna(subset=required).copy()

    if calc_df.empty:
        return calc_df

    calc_df['Macro_Gap'] = calc_df[f'GDP_CAGR_{horizon}Y'] - calc_df[f'ETF_CAGR_{horizon}Y']
    calc_df['R_Macro'] = calc_df['Macro_Gap'].rank(ascending=False).astype(int)
    calc_df['R_REER'] = calc_df['REER_Upside'].rank(ascending=False).astype(int)
    calc_df['R_Bond'] = calc_df['differential with USA'].rank(ascending=False).astype(int)
    calc_df['Currency_Score'] = (calc_df['R_REER'] + calc_df['R_Bond']) / 2
    calc_df['R_Curr'] = calc_df['Currency_Score'].rank().astype(int)
    calc_df['R_Val'] = calc_df['Average Rank'].rank(method="first", ascending=True).astype(int)
    calc_df['R_Nar'] = calc_df['Rank'].rank(method="first", ascending=True).astype(int)
    calc_df['R_Mcap'] = calc_df['Mcap_USD_Bn'].rank(ascending=False).astype(int)
    calc_df['Final_Score'] = (
        calc_df['R_Macro'] * weights["macro"]
        + calc_df['R_Curr'] * weights["curr"]
        + calc_df['R_Val'] * weights["val"]
        + calc_df['R_Nar'] * weights["nar"]
        + calc_df['R_Mcap'] * weights["mcap"]
    )
    calc_df = calc_df.sort_values("Final_Score")
    calc_df['Final Rank'] = range(1, len(calc_df) + 1)
    return calc_df


def excel_value(value):
    return None if pd.isna(value) else value


def _write_dataframe_sheet(wb, title, df):
    ws = wb.create_sheet(title)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append([excel_value(value) for value in row])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 10), 28)
    return ws


def build_editable_model_workbook(
    *,
    master,
    gdp_df,
    etf_df,
    reer_df,
    oil_df,
    mcap_df,
    bond_df,
    val_df,
    nar_df,
    horizon,
    weights,
    terminal_date,
    oil_scenario,
):
    wb = Workbook()
    wb.remove(wb.active)

    _write_dataframe_sheet(wb, "Raw GDP", gdp_df)
    _write_dataframe_sheet(wb, "Raw ETF", etf_df)
    _write_dataframe_sheet(wb, "Raw REER", reer_df)
    _write_dataframe_sheet(wb, "Raw Oil", oil_df)
    _write_dataframe_sheet(wb, "Static Mcap", mcap_df)
    _write_dataframe_sheet(wb, "Static Bond", bond_df)
    _write_dataframe_sheet(wb, "Static Valuation", val_df)
    _write_dataframe_sheet(wb, "Static Narrative", nar_df)
    _write_dataframe_sheet(wb, "Merged Raw Data", master)

    required = [
        f"GDP_CAGR_{horizon}Y",
        f"ETF_CAGR_{horizon}Y",
        "Mcap_USD_Bn",
        "Current_REER",
        "Avg_REER_10Y",
        "10Y bond yield",
        "Average Rank",
        "Rank",
    ]
    model_df = master.dropna(subset=required).copy().sort_values("Country")

    ws = wb.create_sheet("Final Editable Model")
    ws["A1"] = "Editable model export"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "ETF terminal date"
    ws["B2"] = pd.to_datetime(terminal_date).strftime("%Y-%m-%d")
    ws["A3"] = "Horizon"
    ws["B3"] = f"{horizon}Y"
    ws["A4"] = "Oil price move"
    ws["B4"] = oil_scenario
    ws["A6"] = "Weights"
    ws["A6"].font = Font(bold=True)

    weight_rows = [
        ("Macro Gap", weights["macro"]),
        ("Currency", weights["curr"]),
        ("Valuation", weights["val"]),
        ("Narrative", weights["nar"]),
        ("Mcap", weights["mcap"]),
    ]
    for i, (label, value) in enumerate(weight_rows, start=7):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
        ws[f"B{i}"].number_format = "0%"
    ws["A12"] = "Total weight"
    ws["B12"] = "=SUM(B7:B11)"
    ws["B12"].number_format = "0%"
    ws["D7"] = "Edit the weights in B7:B11. Edit valuation and narrative input columns in the table below."
    ws["D7"].alignment = Alignment(wrap_text=True)

    headers = [
        "Final Rank",
        "Country",
        "Region",
        "ETF Ticker",
        f"GDP CAGR {horizon}Y",
        f"ETF CAGR {horizon}Y",
        "Macro Gap",
        "Macro Gap Rank",
        "Current REER",
        "10Y Avg REER",
        "REER Upside",
        "REER Rank",
        "10Y Bond Yield",
        "Bond Diff vs US",
        "Bond Rank",
        "Currency Score",
        "Currency Rank",
        "Valuation Input Rank",
        "Valuation Rank",
        "Narrative Input Rank",
        "Narrative Rank",
        "Mcap USD Bn",
        "Mcap Rank",
        "Final Score",
        "YTD Return",
    ]
    header_row = 15
    data_start = header_row + 1
    data_end = data_start + len(model_df) - 1
    us_rows = model_df.index[model_df["Country"] == "United States"].tolist()
    us_excel_row = data_start + model_df.index.get_loc(us_rows[0]) if us_rows else None

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col_idx, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True)

    for offset, (_, row) in enumerate(model_df.iterrows(), start=0):
        r = data_start + offset
        ws[f"B{r}"] = excel_value(row["Country"])
        ws[f"C{r}"] = excel_value(row["Region"])
        ws[f"D{r}"] = excel_value(row.get("ETF Ticker"))
        ws[f"E{r}"] = excel_value(row[f"GDP_CAGR_{horizon}Y"])
        ws[f"F{r}"] = excel_value(row[f"ETF_CAGR_{horizon}Y"])
        ws[f"G{r}"] = f"=E{r}-F{r}"
        ws[f"H{r}"] = f"=INT(RANK.AVG(G{r},$G${data_start}:$G${data_end},0))"
        ws[f"I{r}"] = excel_value(row["Current_REER"])
        ws[f"J{r}"] = excel_value(row["Avg_REER_10Y"])
        ws[f"K{r}"] = f"=(J{r}-I{r})/J{r}"
        ws[f"L{r}"] = f"=INT(RANK.AVG(K{r},$K${data_start}:$K${data_end},0))"
        ws[f"M{r}"] = excel_value(row["10Y bond yield"])
        ws[f"N{r}"] = f"=M{r}-$M${us_excel_row}" if us_excel_row else f"=M{r}"
        ws[f"O{r}"] = f"=INT(RANK.AVG(N{r},$N${data_start}:$N${data_end},0))"
        ws[f"P{r}"] = f"=(L{r}+O{r})/2"
        ws[f"Q{r}"] = f"=INT(RANK.AVG(P{r},$P${data_start}:$P${data_end},1))"
        ws[f"R{r}"] = excel_value(row["Average Rank"])
        ws[f"S{r}"] = f"=INT(RANK.AVG(R{r},$R${data_start}:$R${data_end},1))"
        ws[f"T{r}"] = excel_value(row["Rank"])
        ws[f"U{r}"] = f"=INT(RANK.AVG(T{r},$T${data_start}:$T${data_end},1))"
        ws[f"V{r}"] = excel_value(row["Mcap_USD_Bn"])
        ws[f"W{r}"] = f"=INT(RANK.AVG(V{r},$V${data_start}:$V${data_end},0))"
        ws[f"X{r}"] = f"=H{r}*$B$7+Q{r}*$B$8+S{r}*$B$9+U{r}*$B$10+W{r}*$B$11"
        ws[f"A{r}"] = f'=RANK.EQ(X{r},$X${data_start}:$X${data_end},1)+COUNTIF($X${data_start}:X{r},X{r})-1'
        ws[f"Y{r}"] = excel_value(row.get("YTD_Return"))

    percent_cols = ["E", "F", "G", "K", "Y"]
    for r in range(data_start, data_end + 1):
        for col in percent_cols:
            ws[f"{col}{r}"].number_format = "0.0%"
        for col in ["I", "J", "M", "N", "P", "X"]:
            ws[f"{col}{r}"].number_format = "0.0"
        ws[f"V{r}"].number_format = "#,##0"

    ws.freeze_panes = f"A{data_start}"
    widths = {
        "A": 11, "B": 18, "C": 16, "D": 11, "E": 12, "F": 12, "G": 12, "H": 12,
        "I": 12, "J": 12, "K": 12, "L": 10, "M": 12, "N": 12, "O": 10,
        "P": 12, "Q": 12, "R": 14, "S": 12, "T": 14, "U": 12, "V": 13,
        "W": 10, "X": 12, "Y": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# --- 4. SIDEBAR CONTROLS ---

with st.sidebar:
    st.header("Settings")
    st.session_state.horizon = st.selectbox("Horizon", options=[1, 3, 5, 10], index=[1, 3, 5, 10].index(st.session_state.horizon))
    term_date = st.date_input("ETF Terminal Date", datetime(2025, 12, 31))
    st.divider()
    
    st.header("Weights")
    st.session_state.w_macro = st.slider("Macro Gap", 0.0, 1.0, st.session_state.w_macro, key="s_macro")
    st.session_state.w_curr = st.slider("Currency Score", 0.0, 1.0, st.session_state.w_curr, key="s_curr")
    st.session_state.w_val = st.slider("Valuation Rank", 0.0, 1.0, st.session_state.w_val, key="s_val")
    st.session_state.w_nar = st.slider("Narrative Rank", 0.0, 1.0, st.session_state.w_nar, key="s_nar")
    st.session_state.w_mcap = st.slider("Mcap Liquidity", 0.0, 1.0, st.session_state.w_mcap, key="s_mcap")
    total_w = round(st.session_state.s_macro + st.session_state.s_curr + st.session_state.s_val + st.session_state.s_nar + st.session_state.s_mcap, 2)
    st.write(f"**Total Weight:** {total_w:.1%}")
    
    st.divider()
    st.header("Scenarios")
    st.session_state.oil_scenario = st.slider("Oil Price Change ($)", 0.0, 50.0, st.session_state.oil_scenario, step=1.0)
    
    st.divider()    
    col1, col2 = st.columns(2)
    if col1.button("Reset Weights"):
        for k, v in DEFAULTS.items(): 
            if k.startswith("w_"): st.session_state[f"s_{k.replace('w_', '')}"] = v
        st.rerun()
    calculate_clicked = col2.button("Calculate Rank", type="primary", disabled=(total_w != 1.0))

# --- 5. DATA ORCHESTRATION ---

try:
    if "valuation_editor_df" not in st.session_state or "narrative_editor_df" not in st.session_state:
        saved_val_df, saved_nar_df = load_research_inputs()
        editor_val_df, editor_nar_df = build_research_editor_inputs(saved_val_df, saved_nar_df)
        st.session_state.valuation_editor_df = editor_val_df
        st.session_state.narrative_editor_df = editor_nar_df

    with st.spinner("Fetching Data..."):
        gdp_j, etf_j, reer_j = fetch_imf_gdp_journey(), fetch_etf_journey(term_date), fetch_bis_reer_journey()
        oil_j = fetch_oil_impact_journey()
        mcap_static = pd.read_csv(STATIC_DIR / "mcap_data.csv")
        bond_df = pd.read_csv(STATIC_DIR / "bond_10y_differentials.csv")
        us_bond_yield = bond_df.loc[bond_df["Country"] == "United States", "10Y bond yield"].iloc[0]
        bond_df["differential with USA"] = bond_df["10Y bond yield"] - us_bond_yield

    strategy_tab, inputs_tab, raw_tab, export_tab = st.tabs([
        "Strategy Output",
        "Editable Inputs",
        "Raw Data & Calculations",
        "Excel Export",
    ])

    with inputs_tab:
        st.subheader("Editable Research Inputs")
        st.caption("Edit final model ranks here. The model output updates only after you click Calculate Rank. Save writes these values back to data/static.")
        st.info("Valuation Rank and Narrative Rank are the final ranks used by the model. The legacy CSV columns are named Average Rank and Rank; those storage names are intentionally hidden here.")
        editor_version = st.session_state.research_editor_version
        editor_col1, editor_col2 = st.columns(2)
        with editor_col1:
            st.markdown("#### Valuation Final Ranks")
            edited_val_df = st.data_editor(
                st.session_state.valuation_editor_df,
                key=f"valuation_editor_{editor_version}",
                use_container_width=True,
                hide_index=True,
                num_rows="fixed",
                disabled=["Country"],
                column_config={
                    "Country": st.column_config.TextColumn("Country"),
                    "Valuation Rank": st.column_config.NumberColumn("Valuation Rank", min_value=1, step=1),
                },
            )
        with editor_col2:
            st.markdown("#### Narrative Final Ranks")
            edited_nar_df = st.data_editor(
                st.session_state.narrative_editor_df,
                key=f"narrative_editor_{editor_version}",
                use_container_width=True,
                hide_index=True,
                num_rows="fixed",
                disabled=["Country"],
                column_config={
                    "Country": st.column_config.TextColumn("Country"),
                    "Narrative Rank": st.column_config.NumberColumn("Narrative Rank", min_value=1, step=1),
                },
            )

        research_errors = validate_research_inputs(edited_val_df, edited_nar_df)
        action_col1, action_col2 = st.columns(2)
        if action_col1.button("Save Research Inputs", disabled=bool(research_errors)):
            saved_val_raw, saved_nar_raw = load_research_inputs()
            save_val_df = merge_research_inputs_for_save(saved_val_raw, edited_val_df, "Valuation Rank", "Average Rank")
            save_nar_df = merge_research_inputs_for_save(saved_nar_raw, edited_nar_df, "Narrative Rank", "Rank")
            save_val_df.to_csv(VALUATION_PATH, index=False)
            save_nar_df.to_csv(NARRATIVE_PATH, index=False)
            editor_val_df, editor_nar_df = build_research_editor_inputs(save_val_df, save_nar_df)
            st.session_state.valuation_editor_df = editor_val_df
            st.session_state.narrative_editor_df = editor_nar_df
            st.session_state.research_editor_version += 1
            st.success("Research inputs saved to data/static.")
            st.rerun()
        if action_col2.button("Reset to Last Saved"):
            saved_val_df, saved_nar_df = load_research_inputs()
            editor_val_df, editor_nar_df = build_research_editor_inputs(saved_val_df, saved_nar_df)
            st.session_state.valuation_editor_df = editor_val_df
            st.session_state.narrative_editor_df = editor_nar_df
            st.session_state.research_editor_version += 1
            st.rerun()
        for error in research_errors:
            st.error(error)

    current_weights = {
        "macro": st.session_state.s_macro,
        "curr": st.session_state.s_curr,
        "val": st.session_state.s_val,
        "nar": st.session_state.s_nar,
        "mcap": st.session_state.s_mcap,
    }

    if calculate_clicked:
        if research_errors:
            st.error("Fix research input errors before calculating rank.")
        else:
            val_df, nar_df = normalize_research_inputs(edited_val_df, edited_nar_df)
            master = build_master_dataset(gdp_j, etf_j, reer_j, oil_j, mcap_static, bond_df, val_df, nar_df)
            calc_df = calculate_strategy(master, st.session_state.horizon, current_weights)
            st.session_state.calc_snapshot = {
                "master": master,
                "calc_df": calc_df,
                "gdp_df": gdp_j,
                "etf_df": etf_j,
                "reer_df": reer_j,
                "oil_df": oil_j,
                "mcap_df": mcap_static,
                "bond_df": bond_df,
                "val_df": val_df,
                "nar_df": nar_df,
                "horizon": st.session_state.horizon,
                "weights": current_weights.copy(),
                "terminal_date": term_date,
                "oil_scenario": st.session_state.oil_scenario,
            }

    snapshot = st.session_state.calc_snapshot

    with export_tab:
        st.subheader("Excel Model Export")
        if snapshot:
            workbook_bytes = build_editable_model_workbook(
                master=snapshot["master"],
                gdp_df=snapshot["gdp_df"],
                etf_df=snapshot["etf_df"],
                reer_df=snapshot["reer_df"],
                oil_df=snapshot["oil_df"],
                mcap_df=snapshot["mcap_df"],
                bond_df=snapshot["bond_df"],
                val_df=snapshot["val_df"],
                nar_df=snapshot["nar_df"],
                horizon=snapshot["horizon"],
                weights=snapshot["weights"],
                terminal_date=snapshot["terminal_date"],
                oil_scenario=snapshot["oil_scenario"],
            )
            st.download_button(
                "Download Editable Excel Model",
                data=workbook_bytes,
                file_name=f"macro_etf_editable_model_{pd.to_datetime(snapshot['terminal_date']).strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Exports the last calculated dashboard snapshot. Recalculate after edits to update this export.",
            )
        else:
            st.info("Calculate rank once to enable the Excel model export.")

    # --- 6. STRATEGY OUTPUT (GATED) ---
    with strategy_tab:
        if not snapshot:
            st.info("💡 Adjust weights in the sidebar and click **Calculate Rank** to generate the strategy output.")
        else:
            h = snapshot["horizon"]
            master = snapshot["master"]
            calc_df = snapshot["calc_df"]
            st.caption("Displayed output and Excel export use the last Calculate Rank snapshot. Recalculate after changing inputs or settings.")
            
            if not calc_df.empty:
                st.header(f"🏆 Strategy Output ({h}Y Horizon)")
                final_map = {
                    'Final Rank': 'Final Rank', 'Country': 'country', 'Region': 'region',
                    'Macro_Gap': 'Macro Gap', 'R_Macro': 'Macro Gap Rank',
                    'Currency_Score': 'Currency Score', 'R_Curr': 'Currency Rank',
                    'R_Val': 'Valuation Rank', 'R_Nar': 'Narrative Rank',
                    'R_Mcap': 'MCap Rank', 'Final_Score': 'Final Score',
                    'YTD_Return': 'YTD $ ETF Return (%)'
                }
                display = calc_df[list(final_map.keys())].rename(columns=final_map)
                pct_cols = ['Macro Gap', 'YTD $ ETF Return (%)']
                for col in pct_cols: display[col] = display[col] * 100
                
                st.dataframe(
                    display.style.format({
                        'Final Rank': '{:d}', 'Final Score': '{:.1f}', 'Currency Score': '{:.1f}',
                        **{c: '{:.1f}' for c in pct_cols},
                        **{c: '{:d}' for c in ['Macro Gap Rank', 'Currency Rank', 'Valuation Rank', 'Narrative Rank', 'MCap Rank']}
                    }).background_gradient(subset=['Final Score', 'Final Rank'], cmap='RdYlGn_r'),
                    use_container_width=True, hide_index=True
                )
            else:
                st.warning("No countries have all required inputs for the selected horizon.")

    with raw_tab:
        st.header("🔍 Raw Data & Calculations")
        t_gdp, t_etf, t_curr, t_qual, t_oil, t_master = st.tabs(["Raw GDP Data", "Raw ETF Data", "Raw Currency Data", "Raw Research Data", "Raw Oil Data", "Merged Data"])
    
        with t_gdp:
            st.subheader("GDP Data: Raw nGDP (bn) -> CAGRs")
            g_cols = ["Country"] + [c for c in gdp_j.columns if "GDP_" in c]
            st.dataframe(gdp_j[g_cols].style.format({c: "{:,.0f}" if "CAGR" not in c else "{:.1%}" for c in g_cols if c != "Country"}), use_container_width=True)
        with t_etf:
            st.subheader("ETF Data: Tickers, Anchor Prices -> CAGRs")
            etf_format = {
                c: "{:.1f}" if c.startswith("P_") else "{:.1%}"
                for c in etf_j.columns
                if c not in ["Country", "ETF Ticker"]
            }
            st.dataframe(etf_j.style.format(etf_format), use_container_width=True)
        with t_curr:
            if snapshot and not calc_df.empty:
                st.subheader("Currency Data: REER & Bond Rank Calculation")
                curr_map = {
                    'Country': 'Country', 'Current_REER': 'Current REER', 'Avg_REER_10Y': '10Y Avg REER',
                    'REER_Upside': 'REER Upside (%)', 'R_REER': 'REER Rank',
                    'differential with USA': 'Bond Diff vs US', 'R_Bond': 'Bond Rank', 'Currency_Score': 'Currency Score (Avg Rank)'
                }
                st.dataframe(calc_df[list(curr_map.keys())].rename(columns=curr_map).style.format({
                    'REER Upside (%)': '{:.1%}', 'REER Rank': '{:d}', 'Bond Rank': '{:d}',
                    'Current REER': '{:.1f}', '10Y Avg REER': '{:.1f}', 'Bond Diff vs US': '{:.1f}', 'Currency Score (Avg Rank)': '{:.1f}'
                }), use_container_width=True)
        with t_qual:
            if snapshot and not calc_df.empty:
                st.subheader("Research Data: Valuation, Narrative & MCap Calculation")
                qual_map = {
                    'Country': 'Country', 'Average Rank': 'Legacy Valuation Storage', 'R_Val': 'Valuation Rank',
                    'R_Nar': 'Narrative Rank', 'Mcap_USD_Bn': '2026 MCap ($ bn.)', 'R_Mcap': 'MCap Rank'
                }
                st.dataframe(calc_df[list(qual_map.keys())].rename(columns=qual_map).style.format({
                    'Legacy Valuation Storage': '{:.1f}', '2026 MCap ($ bn.)': '{:,.0f}',
                    'Valuation Rank': '{:d}', 'Narrative Rank': '{:d}', 'MCap Rank': '{:d}'
                }), use_container_width=True)
                
        with t_oil:
            st.subheader("Crude Oil Impact Proxy")
            oil_cols = ['Country', 'Crude (kg)', 'Crude (mb/d)', 'Conversion_Factor', 'GDP_2025']
            if snapshot and all(col in master.columns for col in oil_cols):
                scenario = snapshot["oil_scenario"]
                oil_calc = master[oil_cols].copy()
                oil_calc['Annual_Impact_USD_Bn'] = (oil_calc['Crude (mb/d)'] * 1e6 * scenario * 365) / 1e9
                oil_calc['GDP_Impact_Pct'] = (oil_calc['Annual_Impact_USD_Bn'] / oil_calc['GDP_2025'])
                
                faulty_countries = oil_calc[ (oil_calc['Country'] == 'Thailand') | (oil_calc['Crude (mb/d)'] > 3.0) & (oil_calc['Country'] != 'China') & (oil_calc['Country'] != 'United States') & (oil_calc['Country'] != 'India') ]
                flagged_list = faulty_countries['Country'].tolist()
                oil_clean = oil_calc[~oil_calc['Country'].isin(flagged_list)].dropna(subset=['Crude (mb/d)', 'GDP_2025'])
                
                oil_view_map = {
                    'Country': 'Country', 'Crude (kg)': '2024 Crude Imports (kg)', 
                    'Conversion_Factor': 'Factor (Barrels/Tonne)', 'Crude (mb/d)': '2024 Imports (mb/d)',
                    'Annual_Impact_USD_Bn': f'${scenario:.0f} Price Move Annual Impact ($ bn)',
                    'GDP_Impact_Pct': 'Impact on 2025 GDP (%)'
                }
                st.dataframe(oil_clean[list(oil_view_map.keys())].rename(columns=oil_view_map).style.format({
                    '2024 Crude Imports (kg)': '{:,.0f}', 'Factor (Barrels/Tonne)': '{:.2f}',
                    '2024 Imports (mb/d)': '{:.1f}', f'${scenario:.0f} Price Move Annual Impact ($ bn)': '{:.1f}',
                    'Impact on 2025 GDP (%)': '{:.1%}'
                }), use_container_width=True)
                if flagged_list:
                    st.info(f"⚠️ **Note:** The following countries were omitted due to unrealistic crude import data (Sanity Check Failure): {', '.join(flagged_list)}")
            else:
                st.warning("Calculate rank to populate the oil impact view.")
                
        with t_master:
            if snapshot:
                st.dataframe(master, use_container_width=True)
            else:
                st.info("Calculate rank to populate the merged model dataset.")

except Exception as e: st.error(f"❌ Error: {e}")
